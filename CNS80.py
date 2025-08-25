import openpyxl
from openpyxl.styles import PatternFill
from difflib import SequenceMatcher

workbook = openpyxl.load_workbook("C:\\File.xlsx") #here You need to paste Your file source
sheet = workbook.active

job_counts = {}
for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2): #here You can change Your column in excel file to another where You have Your positions or another data
    job_title = row[0].value
    if job_title:
        job_counts[job_title] = job_counts.get(job_title, 0) + 1

repeated_jobs = {job: count for job, count in job_counts.items() if count >= 2}
standard_titles = list(repeated_jobs.keys())

red_fill = PatternFill(start_color='B0C4DE', end_color='B0C4DE', fill_type='solid') #here You can change Your color

sheet.cell(row=1, column=3).value = "Suggested position"

def format_title(title):
    if not title:
        return ""
    title = title.lower().title()
    title = title.replace("Bhp", "BHP")
    title = title.replace("General", "General")
    return title

def find_similar_title(title, dictionary, threshold=0.8): #here You can change in how many % the name should be similar 0.8 is 80%, 0.9 is for ex. 90%
    best_match = None
    highest_ratio = 0
    for standard_title in dictionary:
        ratio = SequenceMatcher(None, title.lower(), standard_title.lower()).ratio()
        if ratio > highest_ratio:
            highest_ratio = ratio
            best_match = standard_title
    return best_match if highest_ratio >= threshold else None

for idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=2, max_col=2), start=2):
    job_title = row[0].value
    if job_title:
        if job_title not in repeated_jobs:
            row[0].fill = red_fill
            suggestion = find_similar_title(job_title, standard_titles)
            if suggestion:
                formatted = format_title(suggestion)
                sheet.cell(row=idx, column=3).value = formatted

workbook.save("C:\\File.xlsx")
print("Success")
